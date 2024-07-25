# TODO: (zilong) We need to mark the col names and row names in the table. No need to be markdown format. Can be other formats.
# 1. Remove the surrounding blank cells
# 2. Mark the col names and row names in the same way as Excel (col: A, B, C; row: 1, 2, 3)

import os
import fire
import openpyxl


# DEMO = (
#     'You can write text to a cell in the excel file by calling `excel_read_file` with 2 arguments.\n'
#     '1. The path to the excel file: file_path: str\n'
#     '2. The name of the sheet: sheet_name: str\n'
#     "You can call it by: {'app': 'excel', 'action': 'excel_read_file', 'file_path': ..., 'sheet_name': ...}"
# )

DEMO = (
    "read the excel file to see the existing contents: "
    "{'app': 'excel', 'action': 'read_file', 'file_path': [THE_PATH_TO_THE_EXCEL_FILE]"
)

def construct_action(work_dir, args: dict, py_file_path='/apps/excel_app/excel_read_file.py'):
    # TODO: not sure if we need to specify the file path with the current workdir
    return f'python3 {py_file_path} --file_path {args["file_path"]}'


def excel_read_file(file_path, sheet=None):
    if sheet is None:
        sheet = openpyxl.load_workbook(file_path).active
    else:
        sheet = openpyxl.load_workbook(file_path)[sheet]
    
    content_string = ''
    for row in sheet.iter_rows():
        for cell in row:
            row_idx = cell.row
            col_idx = cell.column
            value = cell.value
            if value is None:
                value = '[Empty Cell]'
            content_string += f'({row_idx}, {col_idx}): {value}\t'
        content_string += '\n'

    return content_string    


def wrap(sheet):
    ob = ("OBSERVATION: The following is the table from the excel file:\n"
         f"{sheet}")
    return ob

def main(file_path, sheet=None, debug=False):
    if not os.path.exists(file_path):
        return f"OBSERVATION: The file {file_path} does not exist. Failed to read the file."

    contents = excel_read_file(file_path, sheet)
    if debug:
        print(contents)
    observation = wrap(contents)
    return observation


if __name__ == '__main__':
    fire.Fire(main)
