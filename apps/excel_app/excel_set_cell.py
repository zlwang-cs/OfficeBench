import os
import fire
import openpyxl


# DEMO = (
#     'You can write text to a cell in the excel file by calling `excel_set_cell` with 5 arguments.\n'
#     '1. The path to the excel file: file_path: str\n'
#     '2. The text to input: text: str\n'
#     '3. The row index of the cell: row_idx: int\n'
#     '4. The column index of the cell: column_idx: int\n'
#     '5. The name of the sheet: sheet_name: str\n'
#     "You can call it by: {'app': 'excel', 'action': 'excel_set_cell', 'file_path': ..., 'text': ..., 'row_idx': ..., 'column_idx': ..., 'sheet_name': ...}"
# )

DEMO = (
    "write text to a cell in the excel file (index starts from 1): "
    "{'app': 'excel', 'action': 'set_cell', 'file_path': [THE_PATH_TO_THE_EXCEL_FILE], 'row_idx': [THE_ROW_INDEX], 'column_idx': [THE_COLUMN_INDEX], 'text': [THE_TEXT_TO_WRITE]}"
)



def construct_action(work_dir, args: dict, py_file_path='/apps/excel_app/excel_set_cell.py'):
    # TODO: not sure if we need to specify the file path with the current workdir
    # return f'python3 {py_file_path} --file_path {args["file_path"]} --text "{args["text"]}" --row_idx {args["row_idx"]} --column_idx {args["column_idx"]}'
    text = str(args["text"])
    if text == 'None':
        text = '[None]'
    return "python3 {} --file_path {} --text '''{}''' --row_idx {} --column_idx {}".format(
        py_file_path, args["file_path"], args["text"], args["row_idx"], args["column_idx"]
    )



def excel_set_cell(file_path, text, row_idx, column_idx, sheet_name=None):
    if text == True:
        text = ''
    try:
        if os.path.exists(file_path):
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()
        
        # choose sheet you want to modify
        if sheet_name is None:
            sheet = workbook.active
        else:
            try:
                sheet = workbook[sheet_name]
            except KeyError:
                # create a new sheet using given name
                sheet = workbook.create_sheet(title=sheet_name)
        
        # Write the text to the specified cell
        row_idx = int(row_idx)
        column_idx = int(column_idx)
        sheet.cell(row=row_idx, column=column_idx, value=text)
        
        # Save the changes to the workbook
        workbook.save(file_path)
        return True
    except Exception as e:
        print(e)
        return False
    

def main(file_path, text, row_idx, column_idx, sheet_name=None):
    if not os.path.exists(file_path):
        return f"OBSERVATION: The file {file_path} does not exist. Failed to write to the file."
    
    print('!!! args:', file_path, text, row_idx, column_idx, sheet_name)

    success = excel_set_cell(file_path, text, row_idx, column_idx, sheet_name)
    if success:
        observation = f'OBSERVATION: Successfully write text to {file_path}'
    else:
        observation = f'OBSERVATION: Failed to write text to {file_path}'
    return observation


if __name__ == '__main__':
    fire.Fire(main)
