import os
import icalendar
import difflib
import sys
import datetime
import openpyxl
import pytz
import glob
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from apps.excel_app import excel_read_file
from apps.word_app import word_read_file
from apps.pdf_app import pdf_read_file
from apps.email_app import email_list_emails
import re
# zip_file is the path to the zipped data for each problem

def _is_number(string):
    try:
        float(string)
        return True
    except ValueError:
        return False

def _evaluate_contain_text(content, args):
    content = content.lower()
    for keyword in args['keywords']:
        keyword = keyword.lower()
        if _is_number(keyword):
            content = content.replace(',', '')
        if keyword in content:
            continue
        else:
            return False
    return True


def evaluate_contain(testbed_dir, args):
    # check if the file contains some keywords
    type = args['doc_type']
    if type == 'email':
        username = args['username']
        email_contents = ''
        if os.path.exists(os.path.join(testbed_dir, 'emails', username)):
            email_contents = email_list_emails.list_emails(username, testbed_dir, -1)
        elif os.path.exists(os.path.join(testbed_dir, 'emails', username.lower())):
            email_contents = email_list_emails.list_emails(username.lower(), testbed_dir, -1)
        else:
            email_accounts = glob.glob(os.path.join(testbed_dir, 'emails', '*'))
            for email_account in os.path.basename(email_account):
                if username.lower() in email_account.lower():
                    email_contents = email_list_emails.list_emails(os.path.basename(email_account), testbed_dir, -1)
                    break
        return _evaluate_contain_text(email_contents, args)
    
    file_path = os.path.join(testbed_dir, args['file'])
    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        print('!!! File not exist:', file_path)
        return False
    if type == 'xlsx':
        helper = excel_read_file.excel_read_file
    elif type == 'txt' or type == 'ics':
        helper = lambda x: open(x).read()
    elif type == 'doc' or type == 'docx':
        helper = word_read_file.word_read_file_into_string
    elif type == 'pdf':
        helper = pdf_read_file.read_pdf_to_string
    else:
        assert False, f'Not implemented doc type: {type}'
    content = helper(file_path)
    return _evaluate_contain_text(content, args)


def evaluate_not_contain(testbed_dir, args):
    return not evaluate_contain(testbed_dir, args)

def evaluate_file_exist(testbed_dir, args):
    file_path = args['file']
    # check if the file exists
    # if file_path in os.listdir(testbed_dir):
    if os.path.exists(os.path.join(testbed_dir, file_path)):
        return True
    else:
        return False 

def evaluate_file_not_exist(testbed_dir, args):
    file_path = args['file']
    # check if the file not exists
    if not os.path.exists(os.path.join(testbed_dir, file_path)):
        return True
    else:
        return False

def _helper_diff_contain_text(input_content, output_content, args):
    if input_content == output_content:
        return False
    else:
        diff = difflib.unified_diff(input_content.split('\n'), output_content.split('\n'), n = 0)
        diff = '\n'.join(list(diff))
        for matches in args['keywords']:
            if matches not in diff:
                return False
    return True

def evaluate_diff_contain_text(testbed_dir, args):
    # check whether the diff of two files contains some keywords
    type = args['doc_type']
    input_file = os.path.join(testbed_dir, args['input_file'])
    output_file = os.path.join(testbed_dir, args['output_file'])
    if type == 'xlsx':
        helper = excel_read_file.excel_read_file
    elif type == 'doc':
        helper = word_read_file.word_read_file
    else:
        assert False, f'Not implemented doc type: {type}'
    input_content = helper(input_file)
    output_content = helper(output_file)
    return _helper_diff_contain_text(input_content, output_content, args)

def evaluate_excel_cell_value(testbed_dir, args):
    file_path = os.path.join(testbed_dir, args['file'])
    if not os.path.exists(file_path):
        print('!!! File not exist:', file_path)
        return False
    content = excel_read_file.excel_read_file(file_path)
    # Match the '(row, col): value' format
    for match in args['matches']:
        pattern = f'({match["row"]}, {match["col"]}): {match["value"]}'
        if pattern in content:
            continue
        else:
            return False
    return True

def evaluate_excel_cell_comparator(testbed_dir, args):
    file_path = os.path.join(testbed_dir, args['file'])
    content = excel_read_file.excel_read_file(file_path)
    # Match the '(row, col): value' format
    for match in args['matches']:
        # regex match pattern (row, col): ***\n
        pattern = '\({}, {}\): (\w+)\t'.format(match["row"], match["col"])
        x = re.search(pattern, content)
        if x:
            value = x.group(1)
            if eval(match['comparator'])(value):
                continue
            else:
                return False
        else:
            return False
    return True

def evaluate_calendar_no_overlap(testbed_dir, args):
    # Go through user's calendar check 
    # whether there is any overlap between events
    username = args['username']
    calendar_file = f'{testbed_dir}/calendar/{username}.ics'
    calendar = icalendar.Calendar.from_ical(open(calendar_file, 'rb').read())
    # sort events by start time

    for x in calendar.subcomponents:
        print(x.get('dtstart').dt)

    utc=pytz.UTC
    def is_naive(dt):
        return dt.tzinfo is None
    def proc_dt(dt):
        if is_naive(dt):
            return utc.localize(dt)
        else:
            return dt
    calendar.subcomponents.sort(key=lambda x: proc_dt(x.get('dtstart').dt))
    # raise ValueError('stop')
    events = []
    for component in calendar.walk():
        if component.name == "VEVENT":
            events.append(component)
    for i in range(len(events)-1):
        if proc_dt(events[i].get('dtend').dt) > proc_dt(events[i+1].get('dtstart').dt):
            event_i_summary = events[i].get('summary')
            event_j_summary = events[i+1].get('summary')
            print(f'Calendar of {username}: Event {event_i_summary} and Event {event_j_summary} overlap')
            return False
    return True

def evaluate_exact_match(testbed_dir, args):
    result_path = os.path.join(testbed_dir, args['result_file'])
    if not os.path.exists(result_path):
        print('!!! File not exist:', result_path)
        return False

    expected_path = os.path.join(testbed_dir, args['expected_file'])
    type = args['doc_type']
    if type != 'xlsx':
        if type == 'txt' or type == 'ics':
            helper = lambda x: open(x).read()
        elif type == 'doc':
            helper = word_read_file.word_read_file
        elif type == 'pdf':
            helper = pdf_read_file.read_pdf
        else:
            assert False, f'Not implemented doc type: {type}'
        result_content = helper(result_path)
        expected_content = helper(expected_path)
        if result_content != expected_content:
            print('!!! Content not match')
            print('result:', result_content)
            print('expected:', expected_content)
            return False
        return True
    else:
        result_sheet = openpyxl.load_workbook(result_path).active
        expected_sheet = openpyxl.load_workbook(expected_path).active

        for row in result_sheet.iter_rows():
            for cell in row:
                row_idx = cell.row
                col_idx = cell.column
                result_value = cell.value
                expected_value = expected_sheet.cell(row=row_idx, column=col_idx).value
                if result_value != expected_value:
                    return False
                
        for row in expected_sheet.iter_rows():
            for cell in row:
                row_idx = cell.row
                col_idx = cell.column
                result_value = result_sheet.cell(row=row_idx, column=col_idx).value
                expected_value = cell.value
                if result_value != expected_value:
                    return False
        
        return True


        
def evaluate(testbed_dir, evaluate_type, args):
    if evaluate_type == 'contain_text':
        return _evaluate_contain_text(testbed_dir, args)
    else:
        raise ValueError('Invalid evaluate type: {}'.format(evaluate_type))

# Example usage
if __name__ == '__main__':
    # test evaluate_calendar_no_overlap
    #print(evaluate_calendar_no_overlap('test_user', {}))
    # test evaluate_excel_diff
    #print(evaluate_excel_diff('/home/zilong-exp/Projects/OfficeBench/OfficeAgentBench/tasks/1-4/testbed/data', {'input_file': 'score.xlsx', 'output_file': 'score_1.xlsx', 'matches': [{'row': 'Bob', 'col': '100'}]}))
    
    # test evaluate_excel_cell_value
    #print(evaluate_excel_cell_value('/home/zilong-exp/Projects/OfficeBench/OfficeAgentBench/tasks/1-4/testbed/data', {'output_file':'score_1.xlsx', 'matches':[{'row': 12, 'col': 2, 'value': '100'}]}))

    # test evaluate_excel_cell_comparator
    def comparator(x):
        # check whether x is a number string
        return x.isdigit() == True and int(x) == 100
    print(evaluate_excel_cell_comparator('/home/zilong-exp/Projects/OfficeBench/OfficeAgentBench/tasks/1-4/testbed/data', {'output_file':'score_1.xlsx', 'matches':[{'row': 12, 'col': 2, 'comparator': comparator}]}))